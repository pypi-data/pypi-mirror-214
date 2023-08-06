# -*- coding: utf-8 -*-
# @Author  : zhousf
# @Function:
import openpyxl
from pathlib import Path
from zhousflib.util import re_util


def unmerge_and_fill_cells(excel_file: Path, merged_cell_rate=2/3, delete_duplicates=True, tmp_excel: Path = None,
                           target_sheet_name=None) -> Path:
    """
    拆分合并单元格并填充有效值
    :param excel_file:
    :param merged_cell_rate: 合并单元格的数量占列数的比例，若大于该比例则拆分该合并单元格
    :param delete_duplicates: 是否对拆分合并单元格的结果去重
    :param tmp_excel: 临时文件，若为空则会更新源文件
    :param target_sheet_name: None第一张表
    :return:
    """
    wb = openpyxl.load_workbook(str(excel_file))
    contain_merge_cells = False
    for sheet_name in wb.sheetnames:
        if target_sheet_name:
            if target_sheet_name != sheet_name:
                continue
        worksheet = wb[sheet_name]
        all_merged_cell_ranges = list(worksheet.merged_cells.ranges)
        rows_deal = {}
        """
        拆分合并单元格
        """
        for merged_cell_range in all_merged_cell_ranges:
            merged_cell = merged_cell_range.start_cell
            worksheet.unmerge_cells(range_string=merged_cell_range.coord)
            start, end = merged_cell_range.coord.split(":")
            start = int(re_util.get_digit_char(start))
            end = int(re_util.get_digit_char(end))
            if (start, end) not in rows_deal:
                rows_deal[(start, end)] = 1
            else:
                rows_deal[(start, end)] += 1

            for row_index, col_index in merged_cell_range.cells:
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.value = merged_cell.value
        """
        找到符合拆分合并单元格条件的单元格rows
        """
        need_fill = []
        for i in rows_deal:
            column_count = rows_deal.get(i)
            # if column_count >= merged_cell_rate * worksheet.max_column:
            #     need_fill.append(i)
            #     contain_merge_cells = True
            need_fill.append(i)
            contain_merge_cells = True
        """
        拆分合并单元格后，对空单元格赋予有效值
        """
        for cells in worksheet.iter_rows():
            for cell in cells:
                row = cell.row
                column = cell.column
                for fill in need_fill:
                    if row == fill[0]:
                        next_cell = worksheet.cell(fill[1], column)
                        if not cell.value and next_cell.value:
                            cell.value = next_cell.value
                        if cell.value and not next_cell.value:
                            next_cell.value = cell.value
        """
        拆分合并单元格后会有重复的两条，这里去重一下
        """
        if len(need_fill) > 0:
            need_fill.sort(key=lambda x: x[0], reverse=False)
        if delete_duplicates:
            # 偏移量，记录删除row的个数
            offset = 0
            for fill in need_fill:
                data = []
                fill = (fill[0]-offset, fill[1]-offset)
                for row_cells in worksheet.iter_rows(min_row=min(fill), max_row=max(fill)):
                    data.append([cell.value for cell in row_cells])
                if len(data) < 2:
                    continue
                first_row_index = min(fill) + 1
                for i in range(1, len(data)):
                    first_row_value = data[i-1]
                    if data[i] == first_row_value:
                        worksheet.delete_rows(idx=first_row_index)
                        offset += 1
    if tmp_excel:
        wb.save(str(tmp_excel))
        wb.close()
        return tmp_excel
    else:
        if contain_merge_cells:
            wb.save(str(excel_file))
        wb.close()
        return excel_file
