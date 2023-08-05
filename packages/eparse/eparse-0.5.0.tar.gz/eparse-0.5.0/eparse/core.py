# -*- coding: utf-8 -*-

'''
excel parser core module
'''

from typing import Dict, List, Tuple

from openpyxl.utils.cell import get_column_letter
import pandas as pd


TableRef = Tuple[int, int, str, str]  # r, c, excel RC, value
TableRec = Tuple[str, str, pd.DataFrame]  # name, excel RC, table


# NOTE: df[n] df.at[r,c] and df.iloc[r,c] are not all the same
#       only with .iloc is it safe to assume index and column
#       names will be ingored ; use iloc when working with
#       sub-tables, e.g. the output from df_parse_table


def df_find_tables(df: pd.DataFrame, loose=False) -> List[TableRef]:
    '''
    finds table corners in a dataframe
    '''

    result = []

    # for each row
    for r in range(df.shape[0]):
        # for each col
        for c in range(df.shape[1]):
            isna_left = True if c == 0 else pd.isna(df.at[r, (c - 1)])
            isna_above = True if r == 0 else pd.isna(df.at[(r - 1), c])
            isna_value = pd.isna(df.at[r, c])

            try:
                isna_right = pd.isna(df.at[r, (c + 1)])
                isna_down = pd.isna(df.at[(r + 1), c])
                isna_corner = pd.isna(df.at[(r + 1), (c + 1)])

                # ensure a 2x2 sparse table
                min_size = any(
                    [
                        not isna_right and not isna_down,
                        not isna_right and not isna_corner,
                        not isna_down and not isna_corner,
                    ]
                )

                if not loose:
                    # ensure a 2x2 dense table
                    min_size = all(
                        [
                            not isna_right,
                            not isna_down,
                            not isna_corner,
                        ]
                    )

            except Exception:
                min_size = False

            if all([isna_left, isna_above, not isna_value, min_size]):
                result.append(
                    (
                        r,
                        c,
                        f'{get_column_letter(c+1)}{r+1}',
                        str(df.at[r, c]),
                    )
                )

    return result


def _is_rowspan(df: pd.DataFrame, r: int, c: int) -> bool:
    '''
    detect a rowspan label
    '''

    try:
        isna_right = pd.isna(df.at[r, (c + 1)])
        isna_down = pd.isna(df.at[(r + 1), c])

        return isna_down and not isna_right

    except KeyError:
        return False


def _has_empty_corner(df: pd.DataFrame, r: int, c: int) -> bool:
    '''
    detect an empty corner
    '''

    try:
        isna_above = pd.isna(df.at[(r - 1), c])
        isna_up_corner = pd.isna(df.at[(r - 1), (c + 1)])

        return isna_above and not isna_up_corner

    except KeyError:
        return False


def df_parse_table(
    df: pd.DataFrame,
    r: int,
    c: int,
) -> pd.DataFrame:
    '''
    extract a table from a dataframe for a given r, c position
    '''

    # make reference adjustments
    if _is_rowspan(df, r, c):
        c += 1

    if _has_empty_corner(df, r, c):
        r -= 1

    _r = r + 1

    # get ending row
    for row in range(r + 1, df.shape[0]):
        if pd.isna(df.at[row, c]):
            break
        _r += 1

    _c = c + 1

    # get ending col
    for col in range(c + 1, df.shape[1]):
        if pd.isna(df.at[r, col]):
            break
        _c += 1

    return df.iloc[r:_r, c:_c]


def df_normalize_data(data: Dict) -> Dict:
    '''
    normalize table data
    '''

    result = {}
    ints = ('row', 'column')
    strs = (
        'value',
        'type',
        'c_header',
        'r_header',
        'excel_RC',
        'name',
        'sheet',
        'f_name',
    )

    for k in ints:
        if k in data:
            result[k] = int(data[k])

    for k in strs:
        if k in data:
            result[k] = str(data[k])

    if 'timestamp' in data:
        if isinstance(data['timestamp'], pd.Timestamp):
            result['timestamp'] = data['timestamp'].to_pydatetime()

    return result


def df_serialize_table(
    df: pd.DataFrame,
    **other_data,
) -> List[Dict]:
    '''
    serialize table into a list of dicts with meta data
    '''

    column_header = df.iloc[0]
    row_header = df.iloc[:, 0]

    result = []

    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            _r = df.index[r]  # excel df row
            _c = df.columns[c]  # excel df col
            result.append(
                df_normalize_data(
                    {
                        'row': r,
                        'column': c,
                        'value': df.iloc[r, c],
                        'type': type(df.iloc[r, c]),
                        'c_header': column_header.iloc[c],
                        'r_header': row_header.iloc[r],
                        'excel_RC': f'{get_column_letter(_c+1)}{_r+1}',
                        **other_data,
                    }
                )
            )

    return result
