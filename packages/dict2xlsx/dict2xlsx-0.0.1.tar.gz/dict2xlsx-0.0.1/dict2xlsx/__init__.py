from openpyxl import Workbook

class Dict2XLSX:
    def __init__(self, data):
        self.book = Workbook()
        sheet = self.book.active

        sheet['A1'] = 'S No.'

        header = {}
        cnt_header = 1
        count = 1

        for obj in data:
            count += 1
            sheet.cell(count, 1).value = count - 1
            for k, v in obj.items():
                if header.get(k, False) == False:
                    sheet.cell(1, cnt_header + 1).value = k
                    header[k] = cnt_header
                    cnt_header += 1
                sheet.cell(count, header[k] + 1).value = v
    
    def save(self, path):
        self.book.save(path)